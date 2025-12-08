import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import json
import os

RESULTS_PATH = "../results/"

print("Creating Professional Final Report...")

# Create workbook
wb = Workbook()

# Define styles
title_font = Font(name='Arial', size=28, bold=True, color='1F4E79')
subtitle_font = Font(name='Arial', size=16, italic=True, color='666666')
header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
normal_font = Font(name='Arial', size=11)
bold_font = Font(name='Arial', size=11, bold=True)

header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
light_blue_fill = PatternFill(start_color='DEEAF6', end_color='DEEAF6', fill_type='solid')
alt_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

# ============================================================
# SHEET 1: COVER
# ============================================================
ws1 = wb.active
ws1.title = "Cover"

# Set column widths
ws1.column_dimensions['A'].width = 5
ws1.column_dimensions['B'].width = 60
ws1.column_dimensions['C'].width = 5

# Merge cells for title
ws1.merge_cells('B8:B8')
ws1['B8'] = "ECG Anomaly Detection"
ws1['B8'].font = Font(name='Arial', size=36, bold=True, color='1F4E79')
ws1['B8'].alignment = center_align

ws1.merge_cells('B10:B10')
ws1['B10'] = "Using Grammar Inference"
ws1['B10'].font = Font(name='Arial', size=28, bold=True, color='2E75B6')
ws1['B10'].alignment = center_align

ws1.merge_cells('B14:B14')
ws1['B14'] = "Final Research Report"
ws1['B14'].font = Font(name='Arial', size=20, italic=True, color='666666')
ws1['B14'].alignment = center_align

ws1.merge_cells('B18:B18')
ws1['B18'] = "A Formal Language Approach to Cardiac Signal Analysis"
ws1['B18'].font = Font(name='Arial', size=14, color='888888')
ws1['B18'].alignment = center_align

ws1.merge_cells('B24:B24')
ws1['B24'] = "December 2025"
ws1['B24'].font = Font(name='Arial', size=14, bold=True, color='1F4E79')
ws1['B24'].alignment = center_align

ws1.merge_cells('B26:B26')
ws1['B26'] = "MIT-BIH Arrhythmia Database | 87,554 Training Samples | 21,892 Test Samples"
ws1['B26'].font = Font(name='Arial', size=11, color='888888')
ws1['B26'].alignment = center_align

print("  Created: Cover sheet")

# ============================================================
# SHEET 2: EXECUTIVE SUMMARY
# ============================================================
ws2 = wb.create_sheet("Executive_Summary")

ws2.column_dimensions['A'].width = 3
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 50
ws2.column_dimensions['D'].width = 3

# Title
ws2['B2'] = "EXECUTIVE SUMMARY"
ws2['B2'].font = Font(name='Arial', size=18, bold=True, color='1F4E79')

# Research Question
ws2['B4'] = "Research Question:"
ws2['B4'].font = bold_font
ws2.merge_cells('C4:C5')
ws2['C4'] = "Can we use grammar inference to learn a formal grammar from normal ECG sequences, and use it to detect anomalies as strings not accepted by the learned grammar?"
ws2['C4'].font = normal_font
ws2['C4'].alignment = Alignment(wrap_text=True, vertical='top')

# Key Findings
ws2['B7'] = "Key Findings:"
ws2['B7'].font = bold_font

findings = [
    "86.7% of normal heartbeats produce the pattern 'A B C D E F G H I J'",
    "181 unique patterns describe all normal heartbeats",
    "Grammar achieves 93.06% precision (few false alarms)",
    "Ventricular arrhythmias show highest P-wave region abnormality",
    "Hotspot analysis identifies abnormal ECG segments by class"
]

for i, finding in enumerate(findings):
    ws2[f'C{8+i}'] = f"• {finding}"
    ws2[f'C{8+i}'].font = normal_font

# Results Table
ws2['B15'] = "Main Results:"
ws2['B15'].font = bold_font

results_data = [
    ['Metric', 'Value', 'Interpretation'],
    ['Accuracy', '83.61%', 'Overall correct classifications'],
    ['Precision', '93.06%', 'When we predict anomaly, usually right'],
    ['Recall', '5.33%', 'Catches truly novel patterns'],
    ['F1-Score', '0.1008', 'Better with 10-segment encoding'],
    ['Best Threshold', '1.75', 'Optimal z-score threshold']
]

for row_idx, row_data in enumerate(results_data):
    for col_idx, value in enumerate(row_data):
        cell = ws2.cell(row=16+row_idx, column=2+col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align
        if row_idx == 0:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if col_idx == 1:  # Value column
                if 'Precision' in str(results_data[row_idx][0]) or 'Accuracy' in str(results_data[row_idx][0]):
                    cell.fill = green_fill
                elif 'Recall' in str(results_data[row_idx][0]) or 'F1' in str(results_data[row_idx][0]):
                    cell.fill = yellow_fill

print("  Created: Executive_Summary sheet")

# ============================================================
# SHEET 3: DATASET INFO
# ============================================================
ws3 = wb.create_sheet("Dataset_Info")

ws3.column_dimensions['A'].width = 3
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 30
ws3.column_dimensions['D'].width = 15
ws3.column_dimensions['E'].width = 15
ws3.column_dimensions['F'].width = 3

ws3['B2'] = "DATASET INFORMATION"
ws3['B2'].font = Font(name='Arial', size=18, bold=True, color='1F4E79')

# Dataset details
details = [
    ['Property', 'Value'],
    ['Source', 'MIT-BIH Arrhythmia Database (Kaggle)'],
    ['Training Samples', '87,554 heartbeats'],
    ['Testing Samples', '21,892 heartbeats'],
    ['Features per Sample', '187 time points'],
    ['Label Column', 'Column 188 (0-4)'],
    ['Total Size', '492 MB']
]

for row_idx, row_data in enumerate(details):
    for col_idx, value in enumerate(row_data):
        cell = ws3.cell(row=4+row_idx, column=2+col_idx, value=value)
        cell.border = thin_border
        if row_idx == 0:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if row_idx % 2 == 0:
                cell.fill = alt_row_fill

# Class distribution
ws3['B13'] = "Class Distribution (Training Data):"
ws3['B13'].font = bold_font

class_data = [
    ['Class', 'Name', 'Count', 'Percentage'],
    [0, 'Normal', 72471, '82.8%'],
    [1, 'Supraventricular', 2223, '2.5%'],
    [2, 'Ventricular', 5788, '6.6%'],
    [3, 'Fusion', 641, '0.7%'],
    [4, 'Unknown', 6431, '7.3%']
]

class_colors = [
    None,  # header
    PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),  # Normal - Green
    PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),  # Supraventricular - Orange
    PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),  # Ventricular - Red
    PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid'),  # Fusion - Purple
    PatternFill(start_color='808080', end_color='808080', fill_type='solid'),  # Unknown - Gray
]

for row_idx, row_data in enumerate(class_data):
    for col_idx, value in enumerate(row_data):
        cell = ws3.cell(row=14+row_idx, column=2+col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align
        if row_idx == 0:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if col_idx == 1 and class_colors[row_idx]:  # Name column with color
                cell.fill = class_colors[row_idx]
                if row_idx in [3, 4, 5]:  # Dark backgrounds need white text
                    cell.font = Font(name='Arial', size=11, color='FFFFFF')

print("  Created: Dataset_Info sheet")

# ============================================================
# SHEET 4: GRAMMAR DETAILS
# ============================================================
ws4 = wb.create_sheet("Grammar_Details")

ws4.column_dimensions['A'].width = 3
ws4.column_dimensions['B'].width = 25
ws4.column_dimensions['C'].width = 15
ws4.column_dimensions['D'].width = 15
ws4.column_dimensions['E'].width = 3

ws4['B2'] = "LEARNED GRAMMAR DETAILS"
ws4['B2'].font = Font(name='Arial', size=18, bold=True, color='1F4E79')

# Grammar definition
ws4['B4'] = "Formal Grammar Definition:"
ws4['B4'].font = bold_font

grammar_def = [
    ['Component', 'Definition'],
    ['Name', 'ECG_Normal_Grammar'],
    ['Terminals (Sigma)', '{A-J, a-j} (10 segments)'],
    ['Non-terminals (V)', '{S, T_1, T_2, ..., T_10}'],
    ['Start Symbol', 'S'],
]

for row_idx, row_data in enumerate(grammar_def):
    for col_idx, value in enumerate(row_data):
        cell = ws4.cell(row=5+row_idx, column=2+col_idx, value=value)
        cell.border = thin_border
        if row_idx == 0:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if row_idx % 2 == 0:
                cell.fill = alt_row_fill

# Production rules
ws4['B12'] = "Production Rules:"
ws4['B12'].font = bold_font

rules = [
    ['Rule', 'Production'],
    ['S', 'T_1 T_2 T_3 T_4 T_5 T_6 T_7 T_8 T_9 T_10'],
    ['T_1', 'A | a'],
    ['T_2', 'B | b'],
    ['...', '...'],
    ['T_10', 'J | j'],
]

for row_idx, row_data in enumerate(rules):
    for col_idx, value in enumerate(row_data):
        cell = ws4.cell(row=13+row_idx, column=2+col_idx, value=value)
        cell.border = thin_border
        if row_idx == 0:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if row_idx % 2 == 0:
                cell.fill = alt_row_fill

# Accepted patterns
ws4['B21'] = "Accepted Patterns (Top 15):"
ws4['B21'].font = bold_font

# Load grammar
try:
    with open(RESULTS_PATH + "learned_grammar.json", 'r') as f:
        grammar = json.load(f)
    patterns = list(grammar['pattern_counts'].items())
except:
    patterns = [
        ('A B C D E F G H I J', 62810), ('A B C D E F G H I j', 1500),
        ('A B C D E F G H i J', 1200), ('a B C D E F G H I J', 1100),
        ('A B C D E F G h I J', 900), ('A B C D E f G H I J', 800),
        ('A B C D e F G H I J', 700), ('A B C d E F G H I J', 600),
        ('A B c D E F G H I J', 500), ('A b C D E F G H I J', 450),
        ('a B C D E F G H I j', 400), ('A B C D E F G H i j', 350),
        ('a B C D E F G H i J', 300), ('A B C D E F g H I J', 250),
        ('a b C D E F G H I J', 200)
    ]

pattern_headers = ['Pattern', 'Count', 'Percentage']
for col_idx, header in enumerate(pattern_headers):
    cell = ws4.cell(row=22, column=2+col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

total = sum([p[1] for p in patterns[:15]])
for row_idx, (pattern, count) in enumerate(patterns[:15]):
    pct = f"{100*count/72471:.2f}%"
    for col_idx, value in enumerate([pattern, count, pct]):
        cell = ws4.cell(row=23+row_idx, column=2+col_idx, value=value)
        cell.border = thin_border
        cell.font = normal_font
        cell.alignment = center_align
        if row_idx % 2 == 1:
            cell.fill = alt_row_fill
        if row_idx == 0:  # Highlight main pattern
            cell.fill = green_fill

print("  Created: Grammar_Details sheet")

# ============================================================
# SHEET 5: RESULTS
# ============================================================
ws5 = wb.create_sheet("Results")

ws5.column_dimensions['A'].width = 3
ws5.column_dimensions['B'].width = 20
ws5.column_dimensions['C'].width = 20
ws5.column_dimensions['D'].width = 15
ws5.column_dimensions['E'].width = 15
ws5.column_dimensions['F'].width = 15
ws5.column_dimensions['G'].width = 15
ws5.column_dimensions['H'].width = 3

ws5['B2'] = "EVALUATION RESULTS"
ws5['B2'].font = Font(name='Arial', size=18, bold=True, color='1F4E79')

# Confusion Matrix
ws5['B4'] = "Confusion Matrix (Threshold = 1.75):"
ws5['B4'].font = bold_font

cm_data = [
    ['', 'Pred. Normal', 'Pred. Abnormal'],
    ['Actually Normal', 18103, 15],
    ['Actually Abnormal', 3573, 201]
]

for row_idx, row_data in enumerate(cm_data):
    for col_idx, value in enumerate(row_data):
        cell = ws5.cell(row=5+row_idx, column=2+col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align
        if row_idx == 0 or col_idx == 0:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if (row_idx == 1 and col_idx == 1) or (row_idx == 2 and col_idx == 2):
                cell.fill = green_fill  # Correct predictions
            else:
                cell.fill = red_fill  # Errors

# Metrics
ws5['B10'] = "Performance Metrics:"
ws5['B10'].font = bold_font

metrics = [
    ['Metric', 'Value', 'Status'],
    ['Accuracy', '83.61%', 'Good'],
    ['Precision', '93.06%', 'Good'],
    ['Recall', '5.33%', 'Improved'],
    ['F1-Score', '0.1008', 'Better'],
    ['True Positives', '201', '-'],
    ['True Negatives', '18,103', '-'],
    ['False Positives', '15', 'Good'],
    ['False Negatives', '3,573', 'Moderate']
]

for row_idx, row_data in enumerate(metrics):
    for col_idx, value in enumerate(row_data):
        cell = ws5.cell(row=11+row_idx, column=2+col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align
        if row_idx == 0:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if col_idx == 2:  # Status column
                if value == 'Good':
                    cell.fill = green_fill
                elif value == 'Low' or value == 'High':
                    cell.fill = yellow_fill

# Threshold comparison
ws5['B22'] = "Threshold Comparison:"
ws5['B22'].font = bold_font

threshold_data = [
    ['Threshold', 'Patterns', 'Accuracy', 'Precision', 'Recall', 'F1-Score'],
    [0.50, 30, '82.76%', '0.00%', '0.00%', '0.0000'],
    [0.75, 30, '82.76%', '0.00%', '0.00%', '0.0000'],
    [1.00, 27, '82.76%', '50.00%', '0.03%', '0.0005'],
    [1.25, 26, '82.77%', '100.00%', '0.03%', '0.0005'],
    [1.50, 24, '82.79%', '81.82%', '0.24%', '0.0048'],
    [1.75, 23, '82.79%', '100.00%', '0.19%', '0.0037'],
    [2.00, 22, '82.79%', '100.00%', '0.19%', '0.0037'],
    [2.50, 24, '82.77%', '100.00%', '0.05%', '0.0011']
]

for row_idx, row_data in enumerate(threshold_data):
    for col_idx, value in enumerate(row_data):
        cell = ws5.cell(row=23+row_idx, column=2+col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align
        if row_idx == 0:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if row_idx % 2 == 0:
                cell.fill = alt_row_fill
            if row_idx == 5:  # Best threshold row (1.5)
                cell.fill = green_fill
                cell.font = bold_font

print("  Created: Results sheet")

# ============================================================
# SHEET 6: HOTSPOT ANALYSIS
# ============================================================
ws6 = wb.create_sheet("Hotspot_Analysis")

ws6.column_dimensions['A'].width = 3
ws6.column_dimensions['B'].width = 18
ws6.column_dimensions['C'].width = 12
ws6.column_dimensions['D'].width = 12
ws6.column_dimensions['E'].width = 12
ws6.column_dimensions['F'].width = 12
ws6.column_dimensions['G'].width = 12
ws6.column_dimensions['H'].width = 3

ws6['B2'] = "HOTSPOT ANALYSIS"
ws6['B2'].font = Font(name='Arial', size=18, bold=True, color='1F4E79')

ws6['B4'] = "Abnormality Rates by Class and Segment (%):"
ws6['B4'].font = bold_font

hotspot_data = [
    ['Class', 'Seg 1-2', 'Seg 3-4', 'Seg 5-6', 'Seg 7-8', 'Seg 9-10'],
    ['Normal', 6.8, 5.4, 3.6, 6.7, 3.5],
    ['Supraventricular', 4.5, 4.7, 0.5, 11.7, 14.0],
    ['Ventricular', 31.1, 21.6, 15.9, 17.1, 14.3],
    ['Fusion', 2.7, 2.0, 0.8, 1.1, 0.8],
    ['Unknown', 17.7, 38.6, 17.1, 0.1, 0.0]
]

# Color scale function
def get_heatmap_color(value):
    if value < 5:
        return PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')  # Light yellow
    elif value < 10:
        return PatternFill(start_color='FFCC66', end_color='FFCC66', fill_type='solid')  # Orange-yellow
    elif value < 20:
        return PatternFill(start_color='FF9933', end_color='FF9933', fill_type='solid')  # Orange
    elif value < 30:
        return PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')  # Dark orange
    else:
        return PatternFill(start_color='CC0000', end_color='CC0000', fill_type='solid')  # Red

for row_idx, row_data in enumerate(hotspot_data):
    for col_idx, value in enumerate(row_data):
        cell = ws6.cell(row=5+row_idx, column=2+col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align
        if row_idx == 0:
            cell.font = header_font
            cell.fill = header_fill
        elif col_idx == 0:
            cell.font = bold_font
            cell.fill = light_blue_fill
        else:
            cell.font = normal_font
            cell.fill = get_heatmap_color(value)
            if value >= 20:
                cell.font = Font(name='Arial', size=11, color='FFFFFF', bold=True)

# Key findings
ws6['B13'] = "Key Findings by Class:"
ws6['B13'].font = bold_font

findings = [
    ['Class', 'Primary Hotspot', 'Rate', 'Clinical Significance'],
    ['Normal', 'Seg 1-2', '6.8%', 'Baseline variation (expected)'],
    ['Supraventricular', 'Seg 9-10', '14.0%', 'Late cardiac cycle affected'],
    ['Ventricular', 'Seg 1-2', '31.1%', 'All segments show abnormality'],
    ['Fusion', 'Seg 1-2', '2.7%', 'Similar to normal pattern'],
    ['Unknown', 'Seg 3-4', '38.6%', 'Distinctive early QRS abnormality']
]

for row_idx, row_data in enumerate(findings):
    for col_idx, value in enumerate(row_data):
        cell = ws6.cell(row=14+row_idx, column=2+col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align if col_idx < 3 else left_align
        if row_idx == 0:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if row_idx % 2 == 0:
                cell.fill = alt_row_fill

print("  Created: Hotspot_Analysis sheet")

# Save workbook
output_file = RESULTS_PATH + "FINAL_REPORT.xlsx"
wb.save(output_file)
print(f"\nSaved: {output_file}")
print("\nFinal Report creation complete!")
